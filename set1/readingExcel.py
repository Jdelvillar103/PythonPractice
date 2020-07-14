import openpyxl, os

os.chdir('C:\\Users\\JOsDP23\\OneDrive\\Desktop') # changes directory to path given
#wb = openpyxl.load_workbook('workbook name')  Opens exsisting workbook
workbook = openpyxl.Workbook() #creates new workbook in current directory
sheet = workbook['Sheet'] 
sheet['A1'] = 42
sheet['A2'] = 'hello'
sheet2 = workbook.create_sheet() # creates new sheet in workbook
sheet2.title = 'NewSheet'
sheet0 = workbook.create_sheet(index = 0,title='Sheet0')
workbook.save('example.xlsx') # Creates and Saves Excel file on current directory
